import streamlit as st
import google.generativeai as genai
import openpyxl
import io
import json
import pandas as pd
import datetime

# --- 1. 網頁基本設定 ---
st.set_page_config(page_title="丸龜採購自動化助手 v2.5", layout="wide")

# --- 2. API Key 與模型初始化 ---
if "GEMINI_API_KEY" in st.secrets:
    genai.configure(api_key=st.secrets["GEMINI_API_KEY"])
    # 使用你剛才測試成功的 2.5 Flash 模型
    # 設定 temperature=0 提高辨識精準度
    model = genai.GenerativeModel(
        model_name='models/gemini-2.5-flash',
        generation_config={
            "temperature": 0,
            "top_p": 0.95,
            "response_mime_type": "application/json",
        }
    )
else:
    st.error("🔑 錯誤：請在 Streamlit Secrets 中設定 GEMINI_API_KEY。")
    st.stop()

st.title("🍣 丸龜採購單自動化系統 (Gemini 2.5 視覺版)")
st.info("本系統直接「看」PDF 圖片辨識內容，可無視任何文字亂碼。")

# 樣板檔案名稱
TEMPLATE_FILE = "丸.xlsx"

# --- 3. 檔案上傳 ---
uploaded_pdfs = st.file_uploader("請上傳 PDF 採購單 (可多選)", type=["pdf"], accept_multiple_files=True)

if st.button("🚀 開始全自動視覺辨識"):
    if uploaded_pdfs:
        try:
            # 讀取 GitHub 上的樣板
            with open(TEMPLATE_FILE, "rb") as f:
                wb = openpyxl.load_workbook(io.BytesIO(f.read()))
            ws = wb.active

            # --- 建立 Excel 座標地圖 ---
            # 日期在第 1 列 (格式: 2026-04-04)
            date_cols = {}
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(row=1, column=c).value)
                if "2026-" in val:
                    date_cols[val[:10]] = c

            # 店號在 B 欄 (從第 3 列開始)
            store_rows = {}
            for r in range(1, ws.max_row + 1):
                val = str(ws.cell(row=r, column=2).value).strip()
                if val and val != "None":
                    store_rows[val] = r
                    # 同時儲存去零後的版本 (例如 001 -> 1) 增加容錯
                    store_rows[val.lstrip('0')] = r

            # --- 處理 PDF ---
            all_extracted_data = []
            
            for pdf_file in uploaded_pdfs:
                st.write(f"🔍 正在掃描檔案：{pdf_file.name}...")
                pdf_bytes = pdf_file.getvalue()

                # 精確的視覺辨識指令
                prompt = """
                你是一個專業的採購單紀錄員。請仔細觀察這份 PDF 每一頁的影像，並提取以下資訊：
                1. 『date』: 尋找「交貨日期」，格式統一為 YYYY-MM-DD。
                2. 『store』: 尋找「交貨地址」欄位中，括號【 】內的前三位店號數字或代碼 (如 052, MT2)。
                3. 『boxes』: 尋找表格中「數量」欄位的數值，只要整數 (如 1.000 變為 1)。

                請嚴格以 JSON 格式回傳列表：
                [{"page": 1, "date": "2026-04-04", "store": "052", "boxes": 1}]
                """

                # 呼叫 Gemini 2.5
                response = model.generate_content([
                    prompt,
                    {"mime_type": "application/pdf", "data": pdf_bytes}
                ])

                try:
                    # 解析 AI 回傳的 JSON
                    page_data = json.loads(response.text)
                    
                    for item in page_data:
                        d = item['date']
                        # 店號補強邏輯：若是純數字且不足三位，自動補零 (如 52 -> 052)
                        s = str(item['store']).strip()
                        if s.isdigit() and len(s) < 3:
                            s = s.zfill(3)
                        
                        b = item['boxes']
                        
                        # 紀錄以便預覽
                        all_extracted_data.append({
                            "檔案": pdf_file.name,
                            "頁碼": item['page'],
                            "日期": d,
                            "店號": s,
                            "箱數": b
                        })

                        # 寫入 Excel
                        col = date_cols.get(d)
                        row = store_rows.get(s)
                        if col and row:
                            ws.cell(row=row, column=col).value = int(b)
                        else:
                            st.warning(f"⚠️ 找不到座標：店號 {s} 在日期 {d} 無對應格子")
                
                except Exception as e:
                    st.error(f"❌ 解析失敗 {pdf_file.name}: {e}")
                    st.code(response.text) # 顯示 AI 回傳內容以便除錯

            # --- 4. 顯示結果預覽 ---
            if all_extracted_data:
                st.subheader("📊 辨識結果預覽")
                st.dataframe(pd.DataFrame(all_extracted_data), use_container_width=True)

                # --- 5. 下載檔案 ---
                output = io.BytesIO()
                wb.save(output)
                st.success("🎉 全數處理完成！")
                st.download_button(
                    label="📩 點我下載填寫好的 Excel",
                    data=output.getvalue(),
                    file_name=f"丸龜自動填寫_{datetime.datetime.now().strftime('%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"系統錯誤: {e}")
    else:
        st.warning("請先上傳 PDF 檔案。")
